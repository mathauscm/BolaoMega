import openpyxl
import warnings

class Apostador:
    def __init__(self, caminho_arquivo):
        self.caminho_arquivo = caminho_arquivo
        # Suprimir warnings temporariamente
        with warnings.catch_warnings(record=True):
            warnings.simplefilter("ignore")
            self.workbook = openpyxl.load_workbook(caminho_arquivo)
        self.sheet = self.workbook.active

    def adicionar(self, nome, dezenas):
        if not self.nome_existe(nome):
            # Encontra a próxima linha disponível na coluna J
            nova_linha_index = self.proxima_linha_disponivel()

            # Adiciona o nome e as dezenas na linha apropriada
            self.sheet.cell(row=nova_linha_index, column=10, value=nome)  # Coluna J (10)
            for i, dezena in enumerate(dezenas):
                self.sheet.cell(row=nova_linha_index, column=11 + i, value=dezena)  # Colunas K a P (11 a 16)

            # Salva o workbook
            self.workbook.save(self.caminho_arquivo)
        else:
            print(f"O nome '{nome}' já existe na tabela. Use a função editar para modificar as dezenas.")

    def proxima_linha_disponivel(self):
        for row in range(3, self.sheet.max_row + 2):  # Começa da linha 3 até a linha máxima + 1
            if self.sheet.cell(row=row, column=10).value is None:
                return row

    def deletar(self, nome):
        for row in self.sheet.iter_rows(min_row=3):  # Começa da linha 3
            if row[9].value and row[9].value.lower() == nome.lower():  # Verifica a coluna J (index 9) e normaliza para minúsculas
                self.sheet.delete_rows(row[0].row)
                self.workbook.save(self.caminho_arquivo)
                break

    def editar(self, nome, novo_nome=None, novas_dezenas=None):
        for row in self.sheet.iter_rows(min_row=3):  # Começa da linha 3
            if row[9].value and row[9].value.lower() == nome.lower():  # Verifica a coluna J (index 9) e normaliza para minúsculas
                if novo_nome:
                    row[9].value = novo_nome
                if novas_dezenas:
                    for i, dezena in enumerate(novas_dezenas):
                        row[10 + i].value = dezena  # Colunas K até P (indexes 10 a 15)
                self.workbook.save(self.caminho_arquivo)
                break

    def nome_existe(self, nome):
        nome = nome.lower()  # Normaliza o nome para minúsculas
        for row in self.sheet.iter_rows(min_row=3):  # Começa da linha 3
            if row[9].value and row[9].value.lower() == nome:  # Verifica a coluna J (index 9) e normaliza para minúsculas
                return True
        return False

# Exemplo de uso
caminho_arquivo = 'C:/Users/Mathaus Carvalho/Documents/Projeto_Nelson/tabela_bolaonelson.xlsx'
apostador_manager = Apostador(caminho_arquivo)

# Receber o nome da pessoa por input
digite_nome = input("Digite o nome do apostador: ")

# Verificar se o nome existe na tabela
existe = apostador_manager.nome_existe(digite_nome)
print(f"O nome '{digite_nome}' existe: {existe}")



"""
# Adicionar um novo apostador, se o nome não existir
if not existe:
    dezenas = list(map(int, input("Digite as dezenas separadas por espaço: ").split()))
    apostador_manager.adicionar(digite_nome, dezenas)
    print(f"O nome '{digite_nome}' foi adicionado com as dezenas {dezenas}.")
else:
    print(f"O nome '{digite_nome}' já existe na tabela.")


# Exemplo de uso
caminho_arquivo = 'C:/Users/Mathaus Carvalho/Documents/Projeto_Nelson/tabela_bolaonelson.xlsx'
apostador_manager = Apostador(caminho_arquivo)

# Adicionar um novo apostador
apostador_manager.adicionar('João', [1, 2, 3, 4, 5, 6])

#Editar apostador
apostador_manager.editar('Maria', novo_nome='Ana', novas_dezenas=[8, 16, 24, 32, 40, 48])

#Deletar apostador
apostador_manager.deletar('Ana')

# Gerar o relatório
apostador_manager.relatorio()



"""



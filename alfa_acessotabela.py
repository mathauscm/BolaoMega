import openpyxl
import warnings
from SCRAPING2 import get_megasena_results

class Apostador:
    def __init__(self, caminho_arquivo):
        self.caminho_arquivo = caminho_arquivo
        with warnings.catch_warnings(record=True):
            warnings.simplefilter("ignore")
            self.workbook = openpyxl.load_workbook(caminho_arquivo)
        self.sheet = self.workbook.active

    def adicionar_resultado(self, data, dezenas):
        nova_linha_index = self.proxima_linha_disponivel()
        self.sheet.cell(row=nova_linha_index, column=1, value=data)  # Coluna A (1)
        for i, dezena in enumerate(dezenas):
            self.sheet.cell(row=nova_linha_index, column=3 + i, value=int(dezena))  # Colunas C a H (3 a 8)
        self.workbook.save(self.caminho_arquivo)
        print(f"Resultado adicionado na linha {nova_linha_index}")

    def proxima_linha_disponivel(self):
        for row in range(3, self.sheet.max_row + 2):
            if self.sheet.cell(row=row, column=1).value is None:
                return row

    def adicionar(self, nome, dezenas):
        if not self.nome_existe(nome):
            nova_linha_index = self.proxima_linha_disponivel()
            self.sheet.cell(row=nova_linha_index, column=10, value=nome)
            for i, dezena in enumerate(dezenas):
                self.sheet.cell(row=nova_linha_index, column=11 + i, value=int(dezena))
            self.workbook.save(self.caminho_arquivo)
        else:
            print(f"O nome '{nome}' já existe na tabela. Use a função editar para modificar as dezenas.")

    def deletar(self, nome):
        for row in self.sheet.iter_rows(min_row=3):
            if row[9].value and row[9].value.lower() == nome.lower():
                self.sheet.delete_rows(row[0].row)
                self.workbook.save(self.caminho_arquivo)
                break

    def editar(self, nome, novo_nome=None, novas_dezenas=None):
        for row in self.sheet.iter_rows(min_row=3):
            if row[9].value and row[9].value.lower() == nome.lower():
                if novo_nome:
                    row[9].value = novo_nome
                if novas_dezenas:
                    for i, dezena in enumerate(novas_dezenas):
                        row[10 + i].value = int(dezena)
                self.workbook.save(self.caminho_arquivo)
                break

    def nome_existe(self, nome):
        nome = nome.lower()
        for row in self.sheet.iter_rows(min_row=3):
            if row[9].value and row[9].value.lower() == nome:
                return True
        return False

    def obter_numeros_sorteados(self):
        numeros_sorteados = []
        for col in range(3, 9):
            for row in range(3, 23):
                valor = self.sheet.cell(row=row, column=col).value
                if valor is not None:
                    numeros_sorteados.append(valor)
        return set(numeros_sorteados)

    def calcular_acertos(self, dezenas, numeros_sorteados):
        return len(set(dezenas) & numeros_sorteados)

    def relatorio(self):
        numeros_sorteados = self.obter_numeros_sorteados()
        print("Relatório de Apostas:")
        for row in self.sheet.iter_rows(min_row=3, min_col=10, max_col=17):
            nome = row[0].value
            dezenas = [cell.value for cell in row[1:7]]
            total_acertos = self.calcular_acertos(dezenas, numeros_sorteados)
            if nome is not None:
                print(f"Nome: {nome}, Total de Acertos: {total_acertos}")

def main():
    caminho_arquivo = 'C:/Users/Mathaus Carvalho/Documents/Projeto_Nelson/tabela_bolaonelson.xlsx'
    apostador_manager = Apostador(caminho_arquivo)

    numeros, concurso_e_data = get_megasena_results()
    print("Números Sorteados no Main:", numeros)
    print("Concurso e data no Main:", concurso_e_data)

    if concurso_e_data and numeros:
        apostador_manager.adicionar_resultado(concurso_e_data, numeros)

    # Gerar o relatório
    apostador_manager.relatorio()

if __name__ == '__main__':
    main()

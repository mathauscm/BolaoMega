import sys
import openpyxl
import warnings
from PyQt5.QtWidgets import (QApplication, QMainWindow, QAction, QFormLayout,
                             QLineEdit, QDialog, QVBoxLayout, QPushButton, QWidget,
                             QLabel, QTableWidget, QTableWidgetItem, QMenuBar, QMessageBox, QFileDialog)
from PyQt5.QtGui import QPixmap
from PyQt5.QtCore import Qt
from SCRAPING2 import get_megasena_results

class Apostador:
    def __init__(self, caminho_arquivo):
        self.caminho_arquivo = caminho_arquivo
        with warnings.catch_warnings(record=True):
            warnings.simplefilter("ignore")
            self.workbook = openpyxl.load_workbook(caminho_arquivo)
        self.sheet = self.workbook.active

    def adicionar_resultado(self, data, dezenas):
        nova_linha_index = self.proxima_linha_disponivel()
        self.sheet.cell(row=nova_linha_index, column=1, value=data)  # Coluna A (1)
        for i, dezena in enumerate(dezenas):
            self.sheet.cell(row=nova_linha_index, column=3 + i, value=int(dezena))  # Colunas C a H (3 a 8)
        self.workbook.save(self.caminho_arquivo)
        print(f"Resultado adicionado na linha {nova_linha_index}")

    def proxima_linha_disponivel(self):
        for row in range(3, self.sheet.max_row + 2):
            if self.sheet.cell(row=row, column=1).value is None:
                return row

    def adicionar(self, nome, dezenas):
        if not self.nome_existe(nome):
            nova_linha_index = self.proxima_linha_disponivel()
            self.sheet.cell(row=nova_linha_index, column=10, value=nome)
            for i, dezena in enumerate(dezenas):
                self.sheet.cell(row=nova_linha_index, column=11 + i, value=int(dezena))
            self.workbook.save(self.caminho_arquivo)
        else:
            print(f"O nome '{nome}' já existe na tabela. Use a função editar para modificar as dezenas.")

    def deletar(self, nome):
        for row in self.sheet.iter_rows(min_row=3):
            if row[9].value and row[9].value.lower() == nome.lower():
                self.sheet.delete_rows(row[0].row)
                self.workbook.save(self.caminho_arquivo)
                break

    def editar(self, nome, novo_nome=None, novas_dezenas=None):
        for row in self.sheet.iter_rows(min_row=3):
            if row[9].value and row[9].value.lower() == nome.lower():
                if novo_nome:
                    row[9].value = novo_nome
                if novas_dezenas:
                    for i, dezena in enumerate(novas_dezenas):
                        row[10 + i].value = int(dezena)
                self.workbook.save(self.caminho_arquivo)
                break

    def nome_existe(self, nome):
        nome = nome.lower()
        for row in self.sheet.iter_rows(min_row=3):
            if row[9].value and row[9].value.lower() == nome:
                return True
        return False

    def obter_numeros_sorteados(self):
        numeros_sorteados = []
        for col in range(3, 9):
            for row in range(3, 23):
                valor = self.sheet.cell(row=row, column=col).value
                if valor is not None:
                    numeros_sorteados.append(valor)
        return set(numeros_sorteados)

    def calcular_acertos(self, dezenas, numeros_sorteados):
        return len(set(dezenas) & numeros_sorteados)

    def relatorio(self):
        numeros_sorteados = self.obter_numeros_sorteados()
        relatorio = "Relatório de Apostas:\n"
        for row in self.sheet.iter_rows(min_row=3, min_col=10, max_col=17):
            nome = row[0].value
            dezenas = [cell.value for cell in row[1:7]]
            total_acertos = self.calcular_acertos(dezenas, numeros_sorteados)
            if nome is not None:
                relatorio += f"Nome: {nome}, Total de Acertos: {total_acertos}\n"
        return relatorio

class InsertDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Inserir Apostador')
        self.setGeometry(100, 100, 300, 200)

        layout = QFormLayout()
        self.name_field = QLineEdit(self)
        self.number_fields = [QLineEdit(self) for _ in range(6)]

        layout.addRow('Nome do Apostador:', self.name_field)
        for i, field in enumerate(self.number_fields, start=1):
            layout.addRow(f'Número {i}:', field)

        self.submit_button = QPushButton('Inserir', self)
        self.submit_button.clicked.connect(self.submit_info)

        main_layout = QVBoxLayout()
        main_layout.addLayout(layout)
        main_layout.addWidget(self.submit_button)
        self.setLayout(main_layout)

    def submit_info(self):
        name = self.name_field.text()
        numbers = [field.text() for field in self.number_fields]
        self.parent().add_apostador(name, numbers)
        self.close()

class EditDialog(QDialog):
    def __init__(self, apostador_manager, parent=None):
        super().__init__(parent)
        self.apostador_manager = apostador_manager
        self.setWindowTitle('Editar Apostador')
        self.setGeometry(100, 100, 300, 300)

        self.name_field = QLineEdit(self)
        self.new_name_field = QLineEdit(self)
        self.number_fields = [QLineEdit(self) for _ in range(6)]

        layout = QFormLayout()
        layout.addRow('Nome do Apostador:', self.name_field)
        layout.addRow('Novo Nome:', self.new_name_field)
        for i, field in enumerate(self.number_fields, start=1):
            layout.addRow(f'Novo Número {i}:', field)

        self.submit_button = QPushButton('Editar', self)
        self.submit_button.clicked.connect(self.submit_info)

        main_layout = QVBoxLayout()
        main_layout.addLayout(layout)
        main_layout.addWidget(self.submit_button)
        self.setLayout(main_layout)

    def submit_info(self):
        name = self.name_field.text()
        new_name = self.new_name_field.text()
        new_numbers = [field.text() for field in self.number_fields if field.text()]
        self.apostador_manager.editar(name, new_name if new_name else None, new_numbers if new_numbers else None)
        self.close()

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.title = 'Sistema de Loteria'
        self.left = 100
        self.top = 100
        self.width = 640
        self.height = 480
        self.apostadores = []
        self.apostador_manager = None
        self.initUI()

    def initUI(self):
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)

        # Configura a imagem de fundo
        self.setBackground()

        mainMenu = self.menuBar()
        fileMenu = mainMenu.addMenu('Arquivo')

        insertAction = QAction('Inserir Apostador', self)
        insertAction.triggered.connect(self.openInsertDialog)
        fileMenu.addAction(insertAction)

        loadExcelAction = QAction('Carregar Arquivo Excel', self)
        loadExcelAction.triggered.connect(self.load_excel)
        fileMenu.addAction(loadExcelAction)

        viewAction = QAction('Ver Relatório', self)
        viewAction.triggered.connect(self.openReportDialog)
        fileMenu.addAction(viewAction)

        scrapeAction = QAction('Obter Resultados da Mega Sena', self)
        scrapeAction.triggered.connect(self.scrape_results)
        fileMenu.addAction(scrapeAction)

        editAction = QAction('Editar Apostador', self)
        editAction.triggered.connect(self.openEditDialog)
        fileMenu.addAction(editAction)

        self.show()

    def setBackground(self):
        # Carrega a imagem e a define como fundo
        background = QPixmap('logo.png')
        backgroundLabel = QLabel(self)
        backgroundLabel.setPixmap(background)
        backgroundLabel.resize(self.width, self.height)
        backgroundLabel.setScaledContents(True)

    def openInsertDialog(self):
        if self.apostador_manager:
            dialog = InsertDialog(self)
            dialog.exec_()
        else:
            self.show_message("Erro", "Por favor, carregue um arquivo Excel primeiro.")

    def openEditDialog(self):
        if self.apostador_manager:
            dialog = EditDialog(self.apostador_manager, self)
            dialog.exec_()
        else:
            self.show_message("Erro", "Por favor, carregue um arquivo Excel primeiro.")

    def openReportDialog(self):
        if self.apostador_manager:
            relatorio = self.apostador_manager.relatorio()
            dialog = ReportDialog(relatorio, self)
            dialog.exec_()
        else:
            self.show_message("Erro", "Por favor, carregue um arquivo Excel primeiro.")

    def scrape_results(self):
        if self.apostador_manager:
            numeros, concurso_e_data = get_megasena_results()
            if numeros and concurso_e_data:
                self.apostador_manager.adicionar_resultado(concurso_e_data, numeros)
                self.show_message("Sucesso", "Resultados da Mega Sena adicionados ao Excel.")
            else:
                self.show_message("Erro", "Falha ao obter resultados da Mega Sena.")
        else:
            self.show_message("Erro", "Por favor, carregue um arquivo Excel primeiro.")

    def load_excel(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        fileName, _ = QFileDialog.getOpenFileName(self, "QFileDialog.getOpenFileName()", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if fileName:
            self.apostador_manager = Apostador(fileName)
            self.show_message("Sucesso", f"Arquivo carregado: {fileName}")

    def add_apostador(self, name, numbers):
        if self.apostador_manager:
            self.apostador_manager.adicionar(name, numbers)
            self.show_message("Sucesso", "Apostador adicionado com sucesso.")
        else:
            self.show_message("Erro", "Por favor, carregue um arquivo Excel primeiro.")

    def show_message(self, title, message):
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Information)
        msgBox.setText(message)
        msgBox.setWindowTitle(title)
        msgBox.setStandardButtons(QMessageBox.Ok)
        msgBox.exec_()

class ReportDialog(QDialog):
    def __init__(self, relatorio, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Relatório de Apostadores')
        self.setGeometry(100, 100, 400, 300)

        layout = QVBoxLayout()
        self.label = QLabel(relatorio, self)
        layout.addWidget(self.label)
        self.setLayout(layout)

def main():
    app = QApplication(sys.argv)
    ex = MainWindow()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
